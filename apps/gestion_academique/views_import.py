# apps/gestion_academique/views_import.py - VERSION CORRIGÉE (Admin + Chef)

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.models import User
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from io import BytesIO

from .models import Etudiant, Departement, Niveau, AnneeAcademique
from .forms import ImportEtudiantsForm


@login_required
def import_etudiants_page(request):
    """Page d'import des étudiants depuis Excel - Admin + Chef"""
    
    # PERMISSION : Admin OU Chef de département
    if not (request.user.profile.is_admin() or request.user.profile.is_chef_departement()):
        messages.error(request, "Seuls l'Administrateur et les Chefs de département peuvent importer des étudiants !")
        return redirect('home')
    
    # Récupérer le département du chef (si c'est un chef)
    departement_limite = None
    if request.user.profile.is_chef_departement():
        departement_limite = request.user.profile.departement
    
    if request.method == 'POST':
        form = ImportEtudiantsForm(request.POST, request.FILES)
        
        if form.is_valid():
            fichier = request.FILES['fichier_excel']
            annee = form.cleaned_data['annee_academique']
            
            # Traiter le fichier avec limitation département si chef
            resultat = traiter_fichier_excel(fichier, annee, departement_limite)
            
            # Afficher les résultats
            if resultat['succes'] > 0:
                messages.success(request, f"✅ {resultat['succes']} étudiant(s) importé(s) avec succès !")
            
            context = {
                'form': ImportEtudiantsForm(),
                'resultat': resultat,
                'departement_limite': departement_limite,
            }
            return render(request, 'gestion_academique/import_etudiants.html', context)
    
    else:
        form = ImportEtudiantsForm()
    
    context = {
        'form': form,
        'departement_limite': departement_limite,
    }
    return render(request, 'gestion_academique/import_etudiants.html', context)


def traiter_fichier_excel(fichier, annee_academique, departement_limite=None):
    """
    Traite un fichier Excel et crée les étudiants
    departement_limite : Si défini, seuls les étudiants de ce département seront importés
    Retourne : {'succes': int, 'erreurs': list, 'details': list}
    """
    
    resultat = {
        'succes': 0,
        'erreurs': [],
        'details': []
    }
    
    try:
        # Lire le fichier Excel
        wb = openpyxl.load_workbook(fichier)
        ws = wb.active
        
        # Récupérer l'en-tête (première ligne)
        headers = [cell.value for cell in ws[1] if cell.value]  # Ignorer les colonnes sans titre
        
        # Vérifier les colonnes obligatoires
        colonnes_requises = ['matricule', 'nom', 'prenom', 'date_naissance', 'lieu_naissance', 'sexe', 'email', 'niveau', 'departement']
        colonnes_manquantes = [col for col in colonnes_requises if col not in headers]
        
        if colonnes_manquantes:
            resultat['erreurs'].append(f"Colonnes manquantes : {', '.join(colonnes_manquantes)}")
            return resultat
        
        # Traiter chaque ligne (à partir de la ligne 2)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Créer un dictionnaire ligne EN GÉRANT LES CELLULES VIDES
                data = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        # Convertir None en chaîne vide
                        data[header] = row[i] if row[i] is not None else ''
                    else:
                        data[header] = ''
                
                # Ignorer les lignes vides (matricule vide)
                matricule = str(data.get('matricule', '')).strip()
                if not matricule:
                    continue
                
                # VÉRIFICATION DÉPARTEMENT (si chef)
                if departement_limite:
                    dept_code = str(data.get('departement', '')).strip().upper()
                    if dept_code != departement_limite.code:
                        resultat['erreurs'].append({
                            'ligne': row_num,
                            'erreur': f"Département '{dept_code}' non autorisé. Vous ne pouvez importer que des étudiants {departement_limite.code}"
                        })
                        continue
                
                # Valider et créer l'étudiant
                etudiant, created = creer_etudiant_depuis_ligne(data, annee_academique, row_num)
                
                if created:
                    resultat['succes'] += 1
                    resultat['details'].append({
                        'ligne': row_num,
                        'matricule': etudiant.matricule,
                        'nom': etudiant.get_full_name(),
                        'departement': etudiant.departement.code,
                        'statut': 'Créé'
                    })
                else:
                    resultat['details'].append({
                        'ligne': row_num,
                        'matricule': matricule,
                        'nom': f"{str(data.get('nom', '')).strip()} {str(data.get('prenom', '')).strip()}",
                        'departement': str(data.get('departement', '')).strip(),
                        'statut': 'Existant (ignoré)'
                    })
            
            except Exception as e:
                resultat['erreurs'].append({
                    'ligne': row_num,
                    'erreur': str(e)
                })
    
    except Exception as e:
        resultat['erreurs'].append({
            'ligne': 'Fichier',
            'erreur': f"Erreur lecture fichier : {str(e)}"
        })
    
    return resultat

def creer_etudiant_depuis_ligne(data, annee_academique, row_num):
    """
    Crée un étudiant depuis une ligne du fichier Excel
    Retourne : (etudiant, created)
    """
    
    try:
        # Vérifier si l'étudiant existe déjà
        matricule = str(data['matricule']).strip()
        
        if Etudiant.objects.filter(matricule=matricule).exists():
            return Etudiant.objects.get(matricule=matricule), False
        
        # Récupérer le département
        dept_code = str(data['departement']).strip().upper()
        try:
            departement = Departement.objects.get(code=dept_code)
        except Departement.DoesNotExist:
            raise ValueError(f"Le département '{dept_code}' n'existe pas")
        
        # Récupérer le niveau
        niveau_code = str(data['niveau']).strip().upper()
        try:
            niveau = Niveau.objects.get(code=niveau_code)
        except Niveau.DoesNotExist:
            raise ValueError(f"Le niveau '{niveau_code}' n'existe pas")
        
        # Parser la date de naissance
        date_naissance = data['date_naissance']
        if isinstance(date_naissance, str):
            # Format attendu : DD/MM/YYYY ou DD-MM-YYYY
            try:
                date_naissance = datetime.strptime(date_naissance, '%d/%m/%Y').date()
            except:
                try:
                    date_naissance = datetime.strptime(date_naissance, '%d-%m-%Y').date()
                except:
                    raise ValueError(f"Format de date invalide : {date_naissance}. Utilisez JJ/MM/AAAA")
        elif isinstance(date_naissance, datetime):
            date_naissance = date_naissance.date()
        
        # Valider le sexe
        sexe = str(data['sexe']).strip().upper()
        if sexe not in ['M', 'F']:
            raise ValueError(f"Sexe invalide : {sexe}. Utilisez M ou F")
        
        # Vérifier si le username (matricule) existe déjà
        if User.objects.filter(username=matricule).exists():
            raise ValueError(f"Le matricule {matricule} est déjà utilisé comme identifiant")
        
        # Créer le compte utilisateur
        try:
            user = User.objects.create_user(
                username=matricule,
                password=matricule,  # Mot de passe = matricule
                first_name=data['prenom'],
                last_name=data['nom']
            )
        except Exception as e:
            if 'UNIQUE constraint failed' in str(e):
                raise ValueError(f"Le matricule {matricule} existe déjà dans le système")
            raise ValueError(f"Erreur lors de la création du compte utilisateur : {str(e)}")
        
        # Générer l'email si vide
        email = data.get('email', '').strip()
        if not email:
            email = f"{matricule.replace('-', '')}@student.uganc.edu.gn"
        
        # Créer l'étudiant
        try:
            etudiant = Etudiant.objects.create(
                matricule=matricule,
                nom=data['nom'],
                prenom=data['prenom'],
                date_naissance=date_naissance,
                lieu_naissance=data['lieu_naissance'],
                sexe=sexe,
                email=email,
                telephone=data.get('telephone', '').strip(),
                departement=departement,
                niveau=niveau,
                annee_academique=annee_academique,
                statut='actif'
            )
        except Exception as e:
            # Supprimer le user créé si l'étudiant ne peut pas être créé
            user.delete()
            
            # Traduire les erreurs SQLite en français
            error_msg = str(e)
            if 'NOT NULL constraint failed' in error_msg:
                if 'email' in error_msg:
                    raise ValueError("L'adresse email est obligatoire")
                elif 'telephone' in error_msg:
                    raise ValueError("Le numéro de téléphone est obligatoire")
                else:
                    raise ValueError("Un champ obligatoire est manquant")
            elif 'UNIQUE constraint failed' in error_msg:
                if 'matricule' in error_msg:
                    raise ValueError(f"Le matricule {matricule} existe déjà")
                elif 'email' in error_msg:
                    raise ValueError(f"L'email {email} est déjà utilisé")
                else:
                    raise ValueError("Une valeur unique est en doublon")
            else:
                raise ValueError(f"Erreur lors de la création de l'étudiant : {error_msg}")
        
        # Lier le profil
        user.profile.role = 'etudiant'
        user.profile.etudiant = etudiant
        user.profile.save()
        
        return etudiant, True
    
    except ValueError as e:
        # Les ValueError sont déjà en français
        raise e
    except Exception as e:
        # Autres erreurs
        raise ValueError(f"Erreur inattendue : {str(e)}")


@login_required
def telecharger_modele_excel(request):
    """Génère et télécharge un fichier Excel modèle"""
    
    # Créer un nouveau workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Étudiants"
    
    # En-têtes
    headers = [
        'matricule', 'nom', 'prenom', 'date_naissance', 'lieu_naissance',
        'sexe', 'email', 'telephone', 'niveau', 'departement'
    ]
    
    # Style pour l'en-tête
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Exemples adaptés selon le rôle
    if request.user.profile.is_chef_departement():
        # Chef : exemples de son département uniquement
        dept_code = request.user.profile.departement.code
        exemples = [
            [f'{111 if dept_code == "NTIC" else 444}-001-234-567', 'Diallo', 'Mamadou', '15/03/2005', 'Conakry', 'M', 'diallo@student.uganc.edu.gn', '+224 123 456 789', 'L1', dept_code],
            [f'{111 if dept_code == "NTIC" else 444}-002-345-678', 'Bah', 'Fatoumata', '20/06/2005', 'Labé', 'F', 'bah@student.uganc.edu.gn', '+224 987 654 321', 'L1', dept_code],
        ]
    else:
        # Admin : exemples des 2 départements
        exemples = [
            ['111-001-234-567', 'Diallo', 'Mamadou', '15/03/2005', 'Conakry', 'M', 'diallo@student.uganc.edu.gn', '+224 123 456 789', 'L1', 'NTIC'],
            ['444-002-345-678', 'Bah', 'Fatoumata', '20/06/2005', 'Labé', 'F', 'bah@student.uganc.edu.gn', '+224 987 654 321', 'L1', 'DL'],
        ]
    
    for row_num, exemple in enumerate(exemples, 2):
        for col_num, valeur in enumerate(exemple, 1):
            ws.cell(row=row_num, column=col_num, value=valeur)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retourner le fichier
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_etudiants.xlsx"'
    
    return response